import cv2
import os
import numpy as np
import face_recognition
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook, load_workbook
from twilio.rest import Client

# Load known faces and their encodings using face_recognition
def load_known_faces(path):
    images = []
    classNames = []
    myList = os.listdir(path)

    for cl in myList:
        curImg = face_recognition.load_image_file(f'{path}/{cl}')
        curEncode = face_recognition.face_encodings(curImg)[0]
        images.append(curEncode)
        classNames.append(os.path.splitext(cl)[0])

    return images, classNames

# Function to create a new sheet with the given title
def create_new_sheet(title):
    wb = Workbook()
    sheet = wb.active
    sheet.title = title
    return wb, sheet

# Function to mark attendance in Excel file
def markAttendance(sheet, name, status, phone_number):
    now = datetime.now()
    dtString = now.strftime('%H:%M:%S')
    if name not in [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row + 1)]:
        new_row = [name, dtString, status, "+" + str(phone_number)]
        sheet.append(new_row)
    else:
        for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
            if row[0].value == name:
                status_cell = row[0].offset(column=2).value
                if status_cell == 'Absent':
                    row[0].offset(column=2).value = status
                    row[0].offset(column=3).value = "+" + str(phone_number)

def get_phone_number_from_csv(name):
    phone_data = pd.read_csv(r'C:\\Users\\Krishna Kumar\\Downloads\\newface\\newface\\phone_numbers.csv')
    row = phone_data[phone_data['Name'] == name]
    if not row.empty:
        phone_number = int(row.iloc[0]['Phone Number'])
        return phone_number
    return ''

# Function to send SMS using Twilio API
def sendSMS(client, to, body):
    try:
        message = client.messages.create(
            to=to,
            from_="+12403926964",
            body=body
        )
        print(f"SMS sent to {to}: {message.sid}")
    except Exception as e:
        print(f"Error sending SMS to {to}: {str(e)}")

# Set up Twilio client
account_sid = 'AC46e63d0ecf4a60f5a3535c5f48355e3e'
auth_token = 'faaefb96a910219e64ab102c79cdde98'
twilio_client = Client(account_sid, auth_token)

# Set the webcam capture duration
capture_duration = timedelta(minutes=1)  # Webcam capture duration of 1 minute

# Initialize webcam capture with OpenCV
cap = cv2.VideoCapture(0)
print(f"Camera opened: {cap.isOpened()}")
if not cap.isOpened():
    print("Error: Couldn't open the camera.")
    exit()

# Load known faces and their encodings
known_face_encodings, known_class_names = load_known_faces(r'C:\\Users\\Krishna Kumar\\Downloads\\newface\\newface\\Training_images')

# Dictionary to keep track of last detection time and status for each person
attendance_records = {name: {'time': datetime.min, 'status': 'Absent'} for name in known_class_names}

# Get today's date for creating a new sheet
today_date = datetime.now().strftime('%Y-%m-%d')

# Check if an Excel file with today's date already exists
excel_file_path = f'C:\\Users\\Krishna Kumar\\Downloads\\newface\\Attendance_{today_date}.xlsx'

# If not, create a new Excel file
if not os.path.exists(excel_file_path):
    wb, today_sheet = create_new_sheet(today_date)
    # Update header in the new sheet
    today_sheet.append(["Name", "Time", "Status", "Phone Number"])
else:
    # If it exists, load the existing Excel file
    wb = load_workbook(excel_file_path)
    today_sheet = wb.active

# Capture webcam video for the defined duration
start_capture_time = datetime.now()
while datetime.now() - start_capture_time < capture_duration:
    success, img = cap.read()
    if not success:
        print("Failed to capture frame. Exiting...")
        break
    if img is None:
        print("Empty frame. Continuing...")
        continue

    # Scale down the input image for faster processing
    small_img = cv2.resize(img, (0, 0), fx=0.25, fy=0.25)

    # Use face_recognition functions
    faces_cur_frame = face_recognition.face_locations(small_img)

    # Process only the largest face if multiple faces are detected
    if faces_cur_frame:
        largest_face_loc = max(faces_cur_frame, key=lambda face: (face[2] - face[0]) * (face[3] - face[1]))
        encode_face = face_recognition.face_encodings(small_img, [largest_face_loc])[0]

        matches = face_recognition.compare_faces(known_face_encodings, encode_face, tolerance=0.6)
        face_dis = face_recognition.face_distance(known_face_encodings, encode_face)

        if any(matches):
            match_index = np.argmin(face_dis)
            name = known_class_names[match_index]

            # Update detection time and status for the matched person
            attendance_records[name]['time'] = datetime.now()
            attendance_records[name]['status'] = 'Present'

            y1, x2, y2, x1 = [coord * 4 for coord in largest_face_loc]
            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
            cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

    # Display the processed frame
    cv2.imshow('Webcam', img)

    # Break the loop if 'q' key is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the webcam and close all windows
cap.release()
cv2.destroyAllWindows()

# Mark absent persons as absent
for name, record in attendance_records.items():
    if record['status'] == 'Absent':
        if (datetime.now() - record['time']) > timedelta(minutes=10):  # Adjust the threshold as needed
            markAttendance(today_sheet, name, 'Absent', get_phone_number_from_csv(name))

# Send the entire attendance report to a single person
# Replace 'your_phone_number' with the actual phone number to receive the report
report_recipient_number = '+916374382550'
attendance_report = ''
for row in today_sheet.iter_rows(min_row=1, max_col=4, max_row=today_sheet.max_row):
    row_data = [cell.value for cell in row]
    attendance_report += '\t'.join(map(str, row_data)) + '\n'

# Send SMS with the entire attendance report
sendSMS(twilio_client, report_recipient_number, f'Today\'s Attendance Report:\n{attendance_report}')

# Save the Excel file
wb.save(excel_file_path)

print("Webcam automation completed.")
